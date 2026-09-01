from io import BytesIO

from openpyxl import load_workbook

from backend.private_export import build_private_xlsx


def test_private_export_contains_base_and_indicator_memory():
    body, filename = build_private_xlsx()
    assert filename.startswith("SEPLAN_BASE_COMPLETA_PRIVADA_")
    assert filename.endswith(".xlsx")
    assert len(body) > 1000

    workbook = load_workbook(BytesIO(body), read_only=True, data_only=False)
    names = set(workbook.sheetnames)
    assert "BASE_COMPLETA" in names
    assert "RESUMO_DASHBOARD" in names
    assert "CONTROLE" in names
    for kpi in range(4, 12):
        assert any(name.startswith(f"KPI{kpi:02d}_") for name in names)

    base = workbook["BASE_COMPLETA"]
    assert base.max_row > 1
    headers = [cell.value for cell in next(base.iter_rows(min_row=1, max_row=1))]
    assert "ProtocoloID" in headers
    assert "ObservacaoAbertura" in headers
    assert "ResponsavelTecnico" in headers
