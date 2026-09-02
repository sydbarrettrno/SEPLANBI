"""Gera a camada pública da carteira de Projetos Públicos sem alterar a fonte XLSX."""
from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "BASE_PROJETOS_PUBLICOS_DASHBOARD.xlsx"
TARGET = ROOT / "data" / "public_projects.json"

FIELD_MAP = {
    "ID": "id",
    "Projeto / Demanda": "project",
    "Grupo": "group",
    "Secretaria / Interface": "interface_raw",
    "Atividade atual": "current_activity",
    "Fase atual": "phase",
    "Status atual": "status",
    "Dependência / Bloqueio": "blocker",
    "Evidência atual da fonte": "evidence",
    "Data de referência": "reference_date",
    "Fonte do universo": "source_universe",
    "Fonte do detalhamento": "source_detail",
    "Confiança": "confidence",
    "Observação de auditoria": "audit_note",
}

INTERFACES = {
    "DESENVOLVIMENTO": "Secretaria de Desenvolvimento",
    "ORDEM PÚBLICA": "Secretaria de Ordem Pública",
    "PESCA E AGRICULTURA": "Secretaria de Pesca e Agricultura",
    "SECULTUR": "Secretaria de Cultura",
    "SEL": "Secretaria de Esporte e Lazer",
    "SEMAI": "Secretaria do Meio Ambiente",
    "SEPLAN": "Secretaria de Planejamento Urbano",
    "SMS": "Saúde / SMS",
    "GABINETE": "Gabinete",
    "GABINETE / CÂMARA": "Gabinete",
    "GABINETE / POLÍCIA CIVIL": "Gabinete",
    "GABINETE / VEREADOR NEI": "Gabinete",
}

GABINETE_INTERFACE = {
    "GABINETE / CÂMARA": "Câmara de Vereadores",
    "GABINETE / VEREADOR NEI": "Câmara de Vereadores",
    "GABINETE / POLÍCIA CIVIL": "Polícia Civil",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["BASE_DASHBOARD"]
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(max_row=1))]
    rows = []
    for cells in sheet.iter_rows(min_row=2):
        source = dict(zip(headers, (cell.value for cell in cells)))
        if clean(source.get("Contabilizar no dashboard")).casefold() != "sim":
            continue
        raw = clean(source.get("Secretaria / Interface"))
        key = raw.upper()
        item = {target: clean(source.get(origin)) for origin, target in FIELD_MAP.items()}
        reference = source.get("Data de referência")
        if isinstance(reference, datetime):
            item["reference_date"] = reference.date().isoformat()
        elif isinstance(reference, date):
            item["reference_date"] = reference.isoformat()
        item["interface"] = INTERFACES.get(key, raw)
        item["gabinete_interface"] = GABINETE_INTERFACE.get(key, "")
        rows.append(item)

    if len(rows) != 20:
        raise RuntimeError(f"Carteira inválida: esperado 20 projetos contabilizáveis, recebido {len(rows)}.")
    if len({row["id"] for row in rows}) != len(rows):
        raise RuntimeError("Carteira inválida: ID de projeto duplicado.")
    TARGET.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"projects": len(rows), "target": str(TARGET)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
