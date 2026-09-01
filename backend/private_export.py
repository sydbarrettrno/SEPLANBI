from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend import analytics, core
from backend.final_entry import dashboard, load_rows, query_from_params
from backend.indicator_views import indicator_view_response
from backend.private_data import load_private_rows


PUBLIC_COLUMNS = [
    "ProtocoloID",
    "NumeroAnoOriginal",
    "ProtocoloAno",
    "DataAbertura",
    "UltimoTramiteDataHora",
    "DataEncerramento",
    "DataSaida",
    "TipoSaida",
    "Macroprocesso",
    "Categoria",
    "StatusOperacional",
    "SetorAtual",
    "GargaloOperacional",
    "DiasSemMovimento",
]
PRIVATE_COLUMNS = [
    "SubassuntoOriginal",
    "ObservacaoAbertura",
    "ObservacaoUltimoTramite",
    "NomeRequerente",
    "ResponsavelTecnico",
    "ResponsavelInterno",
    "PessoaResponsavelExterna",
    "TipoPessoaResponsavel",
    "UsuarioAtualNome",
    "SetorAtualFonte",
    "SituacaoOriginal",
]

HEADER_FILL = PatternFill("solid", fgColor="173E60")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FONT = Font(bold=True)


def _safe(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _export_sources() -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], list[str], list[str]]:
    public_rows = [dict(row) for row in load_rows()]
    private_rows = load_private_rows()
    if not public_rows:
        raise RuntimeError("Base pública sem registros para exportação.")

    public_keys = {key for row in public_rows for key in row.keys()}
    private_keys = {key for row in private_rows.values() for key in row.keys()}

    ordered_public = [column for column in PUBLIC_COLUMNS if column in public_keys]
    ordered_public.extend(sorted(public_keys.difference(ordered_public), key=str.casefold))

    ordered_private = [column for column in PRIVATE_COLUMNS if column in private_keys]
    ordered_private.extend(
        sorted(private_keys.difference(ordered_private).difference({"ProtocoloID"}), key=str.casefold)
    )
    return public_rows, private_rows, ordered_public, ordered_private


def _iter_export_rows(
    public_rows: list[dict[str, Any]],
    private_rows: dict[str, dict[str, str]],
    public_columns: list[str],
    private_columns: list[str],
) -> Iterable[list]:
    for public_row in public_rows:
        protocol_id = str(public_row.get("ProtocoloID") or "").strip()
        private_row = private_rows.get(protocol_id, {})
        yield [
            *[_safe(public_row.get(column)) for column in public_columns],
            *[_safe(private_row.get(column)) for column in private_columns],
        ]


def _style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize(sheet, *, minimum: int = 12, maximum: int = 55) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = minimum
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(maximum, len(value) + 2))
        sheet.column_dimensions[letter].width = width


def _append_key_values(sheet, title: str, values: dict[str, Any]) -> None:
    sheet.append([title, ""])
    row_number = sheet.max_row
    for cell in sheet[row_number]:
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    for key, value in values.items():
        sheet.append([key, _safe(value)])
    sheet.append([])


def _append_table(sheet, title: str, items: list[dict[str, Any]]) -> None:
    if not items:
        return
    headers = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        for key in item.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    if not headers:
        return

    sheet.append([title])
    title_row = sheet.max_row
    sheet[title_row][0].fill = SECTION_FILL
    sheet[title_row][0].font = SECTION_FONT
    sheet.append(headers)
    _style_header(sheet, sheet.max_row)
    for item in items:
        if isinstance(item, dict):
            sheet.append([_safe(item.get(header)) for header in headers])
    sheet.append([])


def _indicator_payload(kpi: int, period_from: str, period_to: str) -> dict[str, Any]:
    params = {
        "kpi": str(kpi),
        "from": period_from,
        "to": period_to,
        "limit": "500",
        "offset": "0",
    }
    return indicator_view_response(params)


def _write_core_indicator_sheet(
    workbook: Workbook,
    kpi: int,
    indicator: str,
    title: str,
    period_from: str,
    period_to: str,
) -> None:
    params = {
        "indicator": indicator,
        "from": period_from,
        "to": period_to,
        "limit": "500",
        "offset": "0",
    }
    query = analytics.query_from_params(params)
    rows = analytics.indicator_rows(query)
    sheet = workbook.create_sheet(f"KPI{kpi:02d}_{title}"[:31])
    sheet.append(["Campo", "Valor"])
    _style_header(sheet)
    sheet.freeze_panes = "A2"

    rules = {
        "received": "Conta protocolos pela DataAbertura dentro do período de referência.",
        "outputs": "Conta saídas operacionais (Concluído ou Encerrado) pela DataSaida dentro do período de referência.",
        "stock": "Retrato atual dos protocolos não terminais na data de corte; from/to não reconstrói estoque histórico.",
    }
    _append_key_values(sheet, "IDENTIFICAÇÃO", {
        "Indicador": kpi,
        "Nome": title.replace("_", " ").title(),
        "Universo": indicator,
        "Regra": rules[indicator],
        "Total": len(rows),
        "Período": f"{period_from} a {period_to}" if indicator != "stock" else f"Posição em {period_to}",
    })

    totals = analytics.control_totals(query)
    _append_key_values(sheet, "CONTROLE / RECONCILIAÇÃO", totals)

    if indicator in {"received", "outputs"}:
        _append_key_values(sheet, "COMPARAÇÃO HOMÓLOGA", analytics.indicator_comparison(query, indicator))

    if indicator == "outputs":
        _append_table(sheet, "FLUXO MENSAL — ENTRADAS X SAÍDAS", analytics.monthly_flow(query))

    for dimension in analytics.DEFAULT_GROUPS[indicator]:
        grouped = analytics.group_rows(rows, (dimension,), indicator)
        normalized = [
            {
                "Dimensão": dimension,
                "Item": next(iter(item.get("keys", {}).values()), "Não informado"),
                "Quantidade": item.get("value", 0),
            }
            for item in grouped
        ]
        _append_table(sheet, f"AGRUPAMENTO — {analytics.DIMENSION_LABELS[dimension].upper()}", normalized)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autosize(sheet)


def _write_indicator_sheet(workbook: Workbook, kpi: int, payload: dict[str, Any]) -> None:
    name = str(payload.get("name") or f"Indicador {kpi:02d}")
    sheet = workbook.create_sheet(f"KPI{kpi:02d}_{name}"[:31])
    sheet.append(["Campo", "Valor"])
    _style_header(sheet)
    sheet.freeze_panes = "A2"

    coverage = payload.get("coverage")
    limitation = coverage.get("limitation") if isinstance(coverage, dict) else None
    _append_key_values(sheet, "IDENTIFICAÇÃO", {
        "Indicador": kpi,
        "Nome": name,
        "Status": payload.get("status"),
        "Regra": payload.get("rule"),
        "Motivo/limitação": payload.get("reason") or limitation,
    })

    metrics = payload.get("metrics")
    if isinstance(metrics, dict):
        _append_key_values(sheet, "RESULTADO / MEMÓRIA DE CÁLCULO", metrics)
    elif metrics is None:
        _append_key_values(sheet, "RESULTADO / MEMÓRIA DE CÁLCULO", {"Resultado": "Não calculado / não homologado"})

    for section_key, section_title in (
        ("comparison", "COMPARAÇÃO"),
        ("cohort", "COORTE"),
        ("coverage", "COBERTURA"),
        ("context", "CONTEXTO"),
    ):
        section = payload.get(section_key)
        if isinstance(section, dict):
            _append_key_values(sheet, section_title, section)

    for key, title in (
        ("monthly", "SÉRIE MENSAL"),
        ("categories", "CATEGORIAS"),
        ("bands", "FAIXAS"),
        ("distribution", "DISTRIBUIÇÃO"),
        ("sectors", "SETORES"),
        ("responsibilities", "RESPONSABILIDADES"),
        ("stages", "ETAPAS"),
        ("sla_rules", "REGRAS DE PRAZO"),
    ):
        items = payload.get(key)
        if isinstance(items, list):
            _append_table(sheet, title, items)

    records = payload.get("records")
    if isinstance(records, dict) and isinstance(records.get("items"), list):
        _append_table(sheet, "REGISTROS QUE COMPÕEM O INDICADOR", records["items"])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autosize(sheet)


def _write_overview_sheet(workbook: Workbook, period_from: str, period_to: str) -> None:
    payload = dashboard(query_from_params({"from": period_from, "to": period_to, "limit": "1", "offset": "0"}))
    sheet = workbook.create_sheet("RESUMO_DASHBOARD")
    sheet.append(["Elemento", "Valor", "Regra/observação"])
    _style_header(sheet)
    metrics = payload.get("metrics", {})
    turnaround = metrics.get("turnaround", {}) if isinstance(metrics, dict) else {}
    rows = [
        ("Recebidos", metrics.get("received"), "Protocolos recebidos no período padrão do dashboard."),
        ("Finalizados", metrics.get("concluded"), "Saídas operacionais no período padrão do dashboard."),
        ("Saldo do período", metrics.get("period_balance"), "Recebidos menos finalizados."),
        ("Estoque atual", metrics.get("stock"), "Protocolos não terminais na data de corte."),
        ("Fila interna", metrics.get("internal_queue"), "Parcela do estoque atribuída à fila interna SEPLAN."),
        ("Aguardando externo", metrics.get("external_wait"), "Parcela do estoque aguardando responsável externo."),
        ("Tempo mediano", turnaround.get("median_days"), "Mediana de abertura até saída operacional."),
        ("Tempo médio", turnaround.get("mean_days"), "Média de abertura até saída operacional."),
        ("Tempo P90", turnaround.get("p90_days"), "Percentil 90 do tempo de tramitação."),
    ]
    for row in rows:
        sheet.append(row)
    sheet.append([])
    sheet.append(["Período de referência", f"{period_from} a {period_to}", "Recorte padrão da versão exportada."])
    sheet.append(["Data de corte da base", core.metadata().get("source_updated_at", ""), "Última versão carregada no sistema."])

    charts = payload.get("charts", {})
    if isinstance(charts, dict):
        flow = charts.get("flow")
        if isinstance(flow, list):
            _append_table(sheet, "DADOS DO GRÁFICO — FLUXO MENSAL", flow)

    _autosize(sheet)


def build_private_xlsx() -> tuple[bytes, str]:
    public_rows, private_rows, public_columns, private_columns = _export_sources()
    rows = list(_iter_export_rows(public_rows, private_rows, public_columns, private_columns))
    if not rows:
        raise RuntimeError("Base privada sem registros para exportação.")

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "BASE_COMPLETA"
    headers = [*public_columns, *private_columns]
    sheet.append(headers)
    _style_header(sheet)

    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "ProtocoloID": 18,
        "NumeroAnoOriginal": 18,
        "ProtocoloAno": 12,
        "DataAbertura": 14,
        "UltimoTramiteDataHora": 22,
        "DataEncerramento": 18,
        "DataSaida": 14,
        "TipoSaida": 16,
        "Macroprocesso": 28,
        "Categoria": 30,
        "StatusOperacional": 30,
        "SetorAtual": 32,
        "GargaloOperacional": 26,
        "DiasSemMovimento": 18,
        "SubassuntoOriginal": 28,
        "ObservacaoAbertura": 55,
        "ObservacaoUltimoTramite": 55,
        "NomeRequerente": 34,
        "ResponsavelTecnico": 34,
        "ResponsavelInterno": 30,
        "PessoaResponsavelExterna": 34,
        "TipoPessoaResponsavel": 24,
        "UsuarioAtualNome": 28,
        "SetorAtualFonte": 32,
        "SituacaoOriginal": 22,
    }
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)

    for field in ("ObservacaoAbertura", "ObservacaoUltimoTramite"):
        if field in headers:
            field_index = headers.index(field)
            for row in sheet.iter_rows(min_row=2):
                row[field_index].alignment = Alignment(vertical="top", wrap_text=True)

    meta = core.metadata()
    default_period = meta.get("default_period", {}) if isinstance(meta.get("default_period"), dict) else {}
    source_date = str(meta.get("source_updated_at") or "")[:10]
    period_from = str(default_period.get("from") or (f"{source_date[:4]}-01-01" if source_date else "2026-01-01"))
    period_to = str(default_period.get("to") or source_date)
    if not period_to:
        raise RuntimeError("Data de referência da base indisponível para cálculo dos indicadores.")

    _write_overview_sheet(workbook, period_from, period_to)
    _write_core_indicator_sheet(workbook, 1, "received", "RECEBIDOS", period_from, period_to)
    _write_core_indicator_sheet(workbook, 2, "outputs", "FINALIZADOS", period_from, period_to)
    _write_core_indicator_sheet(workbook, 3, "stock", "ESTOQUE", period_from, period_to)
    for kpi in range(4, 12):
        _write_indicator_sheet(workbook, kpi, _indicator_payload(kpi, period_from, period_to))

    metadata_sheet = workbook.create_sheet("CONTROLE")
    metadata_sheet.append(["Campo", "Valor"])
    metadata_sheet.append(["Data de referência", meta.get("source_updated_at", "")])
    metadata_sheet.append(["Período dos cálculos", f"{period_from} a {period_to}"])
    metadata_sheet.append(["Protocolos exportados", len(rows)])
    metadata_sheet.append(["Colunas públicas exportadas", len(public_columns)])
    metadata_sheet.append(["Colunas privadas exportadas", len(private_columns)])
    metadata_sheet.append(["Taxonomia", meta.get("taxonomy_version", "V07")])
    metadata_sheet.append(["Classificação", "BASE PRIVADA — USO INTERNO"])
    metadata_sheet.append(["Conteúdo", "Inclui todos os campos disponíveis na base pública canônica, todos os campos autorizados da camada privada, resumo executivo e memória de cálculo dos 11 indicadores."])
    metadata_sheet.append(["Segurança", "Exportação autorizada no backend; não publicar este arquivo em GitHub/Vercel como artefato estático."])
    _style_header(metadata_sheet)
    metadata_sheet.column_dimensions["A"].width = 28
    metadata_sheet.column_dimensions["B"].width = 100
    for row in metadata_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    filename_date = source_date.replace("-", "") or "atual"
    filename = f"SEPLAN_BASE_COMPLETA_PRIVADA_{filename_date}.xlsx"
    return output.getvalue(), filename
