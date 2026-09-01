from io import BytesIO
import unittest

from openpyxl import Workbook

from backend.private_import import parse_private_xlsx


def _workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class PrivateImportTests(unittest.TestCase):
    def test_complete_base_maps_authorized_fields_and_ignores_tax_ids(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BASE_COMPLETA"
        sheet.append([
            "Número/Ano", "Requerente - Nome Razão", "Observação Abertura", "Último Trâmite - Observação",
            "ProtocoloID_Sistema", "EscopoSistema", "Responsável - Nome", "Usuário Atual - Nome",
            "Situação", "Subassunto - Descrição", "Centro de Custo Atual - Descrição",
            "Requerente - CPF/CNPJ", "Responsável - CPF/CNPJ",
        ])
        sheet.append([
            "43781/2026", "Pessoa Requerente", "Pedido inicial", "Último trâmite",
            "2026-43781", "INCLUIDO_2025_MAIS", "RT Teste", "Servidor Teste",
            "Aberto", "ALVARÁ", "Engenharia", "111.111.111-11", "222.222.222-22",
        ])

        source, records = parse_private_xlsx(_workbook_bytes(workbook))
        self.assertEqual(source, "BASE_COMPLETA")
        record = records["2026-43781"]
        self.assertEqual(record["NomeRequerente"], "Pessoa Requerente")
        self.assertEqual(record["ResponsavelTecnico"], "RT Teste")
        self.assertEqual(record["ResponsavelInterno"], "Servidor Teste")
        self.assertEqual(record["ObservacaoAbertura"], "Pedido inicial")
        self.assertEqual(record["ObservacaoUltimoTramite"], "Último trâmite")
        self.assertEqual(record["SetorAtualFonte"], "Engenharia")
        self.assertNotIn("Requerente - CPF/CNPJ", record)
        self.assertNotIn("Responsável - CPF/CNPJ", record)

    def test_raw_base_maps_authorized_fields_and_ignores_tax_ids(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BASE23-26"
        sheet.append([
            "Número/Ano", "Requerente", "Categoria", "ObsAbertura", "UltTramiteOBS",
            "Situação", "CPF_CNPJ_REQUERENTE", "NomeRT", "CPF_CNPJ_RT", "CCAtual", "UsuarioAtual",
        ])
        sheet.append([
            "43781/2026", "Pessoa Requerente", "ALVARÁ", "Pedido inicial", "Último trâmite",
            "Aberto", "111.111.111-11", "RT Teste", "222.222.222-22", "Engenharia", "Servidor Teste",
        ])

        source, records = parse_private_xlsx(_workbook_bytes(workbook))
        self.assertEqual(source, "BASE23-26")
        record = records["2026-43781"]
        self.assertEqual(record["NomeRequerente"], "Pessoa Requerente")
        self.assertEqual(record["ResponsavelTecnico"], "RT Teste")
        self.assertEqual(record["ResponsavelInterno"], "Servidor Teste")
        self.assertEqual(record["ObservacaoAbertura"], "Pedido inicial")
        self.assertEqual(record["ObservacaoUltimoTramite"], "Último trâmite")
        self.assertEqual(record["UsuarioAtualNome"], "Servidor Teste")
        self.assertNotIn("CPF_CNPJ_REQUERENTE", record)
        self.assertNotIn("CPF_CNPJ_RT", record)

    def test_indicator_base_maps_received_and_internal_responsible(self):
        workbook = Workbook()
        received = workbook.active
        received.title = "01_RECEBIDOS"
        received.append([
            "ProtocoloID", "ResponsavelTecnico", "ObservacaoUltimoTramite",
            "NomeRequerente", "PessoaResponsavelExterna", "TipoPessoaResponsavel",
        ])
        received.append([
            "2026-43781", "RT Teste", "Último trâmite", "Pessoa Requerente", "RT Teste", "Responsável Técnico",
        ])
        stock = workbook.create_sheet("03_ESTOQUE")
        stock.append(["ProtocoloID", "ResponsavelInterno"])
        stock.append(["2026-43781", "Servidor Teste"])

        source, records = parse_private_xlsx(_workbook_bytes(workbook))
        self.assertEqual(source, "BASE_INDICADORES")
        record = records["2026-43781"]
        self.assertEqual(record["ResponsavelInterno"], "Servidor Teste")
        self.assertEqual(record["ResponsavelTecnico"], "RT Teste")
        self.assertEqual(record["NomeRequerente"], "Pessoa Requerente")

    def test_rejects_unrelated_workbook(self):
        workbook = Workbook()
        workbook.active.title = "PLANILHA_ALEATORIA"
        workbook.active.append(["A", "B"])
        with self.assertRaises(ValueError):
            parse_private_xlsx(_workbook_bytes(workbook))


if __name__ == "__main__":
    unittest.main()
