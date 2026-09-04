from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
import lzma
import os
from typing import Any

from openpyxl import load_workbook

from backend.final_entry import load_rows
from backend.private_data import BLOB_PATH, PRIVATE_FIELDS_V3, load_private_rows


MIN_PUBLIC_COVERAGE = 0.95


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _protocol_id(value: Any) -> str:
    text = _clean(value).replace("/", "-")
    parts = [part.strip() for part in text.split("-") if part.strip()]
    if len(parts) != 2 or not all(part.isdigit() for part in parts):
        return ""
    first = int(parts[0])
    second = int(parts[1])
    if 2020 <= first <= 2100:
        year, number = first, second
    elif 2020 <= second <= 2100:
        year, number = second, first
    else:
        return ""
    return f"{year}-{number}"


def _blank_record(protocol_id: str) -> dict[str, str]:
    record = {field: "" for field in PRIVATE_FIELDS_V3}
    record["ProtocoloID"] = protocol_id
    return record


def _find_header(sheet, required: set[str]) -> tuple[list[str], int]:
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        headers = [_clean(value) for value in row]
        if required.issubset(set(headers)):
            return headers, index
    raise ValueError(f"Cabeçalho esperado não encontrado na aba {sheet.title}.")


def _sheet_records(sheet, required: set[str]):
    headers, header_row = _find_header(sheet, required)
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(_clean(value) for value in values):
            continue
        yield {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }


def _parse_raw_base(sheet) -> dict[str, dict[str, str]]:
    records: dict[str, dict[str, str]] = {}
    required = {"Número/Ano", "Requerente", "ObsAbertura", "UltTramiteOBS"}
    for source in _sheet_records(sheet, required):
        protocol_id = _protocol_id(source.get("Número/Ano"))
        if not protocol_id:
            continue
        year = int(protocol_id.split("-", 1)[0])
        if year < 2025:
            continue
        if protocol_id in records:
            raise ValueError("A planilha possui protocolos duplicados.")

        requester = _clean(source.get("Requerente"))
        technical = _clean(source.get("NomeRT"))
        internal = _clean(source.get("UsuarioAtual"))
        record = _blank_record(protocol_id)
        record.update({
            "NomeRequerente": requester,
            "ResponsavelTecnico": technical,
            "ResponsavelInterno": internal,
            "ObservacaoAbertura": _clean(source.get("ObsAbertura")),
            "ObservacaoUltimoTramite": _clean(source.get("UltTramiteOBS")),
            "SubassuntoOriginal": _clean(source.get("Categoria")),
            "UsuarioAtualNome": internal,
            "SituacaoOriginal": _clean(source.get("Situação")),
            "SetorAtualFonte": _clean(source.get("CCAtual")),
            "PessoaResponsavelExterna": technical or requester,
            "TipoPessoaResponsavel": "Responsável/RT" if technical else ("Requerente" if requester else ""),
        })
        records[protocol_id] = record
    return records


def _parse_ipm_report(sheet) -> dict[str, dict[str, str]]:
    """Extrai a camada privada diretamente do relatório IPM atual (aba Report).

    Esta é a fonte usada na atualização incremental do dashboard. CPF/CNPJ nunca
    é copiado para o artefato privado; somente os campos previstos no contrato V3.
    """
    records: dict[str, dict[str, str]] = {}
    required = {
        "Número/Ano",
        "Requerente - Nome Razão",
        "Observação Abertura",
        "Último Trâmite - Observação",
        "Subassunto - Descrição",
    }
    for source in _sheet_records(sheet, required):
        protocol_id = _protocol_id(source.get("Número/Ano"))
        if not protocol_id:
            continue
        year = int(protocol_id.split("-", 1)[0])
        if year < 2025:
            continue
        if protocol_id in records:
            raise ValueError("A planilha possui protocolos duplicados no escopo 2025+.")

        requester = _clean(source.get("Requerente - Nome Razão"))
        technical = _clean(source.get("Responsável - Nome"))
        internal_name = _clean(source.get("Usuário Atual - Nome"))
        internal_code = _clean(source.get("Usuário Atual - Código"))
        internal = internal_name or internal_code
        record = _blank_record(protocol_id)
        record.update({
            "NomeRequerente": requester,
            "ResponsavelTecnico": technical,
            "ResponsavelInterno": internal,
            "ObservacaoAbertura": _clean(source.get("Observação Abertura")),
            "ObservacaoUltimoTramite": _clean(source.get("Último Trâmite - Observação")),
            "SubassuntoOriginal": _clean(source.get("Subassunto - Descrição")),
            "UsuarioAtualNome": internal_name or internal_code,
            "SituacaoOriginal": _clean(source.get("Situação")),
            "SetorAtualFonte": _clean(source.get("Centro de Custo Atual - Descrição")),
            "PessoaResponsavelExterna": technical or requester,
            "TipoPessoaResponsavel": "Responsável/RT" if technical else ("Requerente" if requester else ""),
        })
        records[protocol_id] = record
    return records


def _parse_complete_base(sheet) -> dict[str, dict[str, str]]:
    """Extrai somente a camada privada autorizada da BASE_COMPLETA.

    A planilha completa contém CPF/CNPJ, mas esses campos não são copiados para o
    artefato privado do sistema. Somente protocolos canônicos do escopo 2025+ são
    considerados, usando ProtocoloID_Sistema quando disponível.
    """
    records: dict[str, dict[str, str]] = {}
    required = {
        "Número/Ano",
        "Requerente - Nome Razão",
        "Observação Abertura",
        "Último Trâmite - Observação",
        "ProtocoloID_Sistema",
    }
    for source in _sheet_records(sheet, required):
        protocol_id = _protocol_id(source.get("ProtocoloID_Sistema")) or _protocol_id(source.get("Número/Ano"))
        if not protocol_id:
            continue
        year = int(protocol_id.split("-", 1)[0])
        if year < 2025:
            continue

        scope = _clean(source.get("EscopoSistema"))
        if scope and scope != "INCLUIDO_2025_MAIS":
            continue
        if protocol_id in records:
            raise ValueError("A planilha possui protocolos duplicados no escopo do sistema.")

        requester = _clean(source.get("Requerente - Nome Razão"))
        technical = _clean(source.get("Responsável - Nome"))
        internal = _clean(source.get("Usuário Atual - Nome"))
        record = _blank_record(protocol_id)
        record.update({
            "NomeRequerente": requester,
            "ResponsavelTecnico": technical,
            "ResponsavelInterno": internal,
            "ObservacaoAbertura": _clean(source.get("Observação Abertura")),
            "ObservacaoUltimoTramite": _clean(source.get("Último Trâmite - Observação")),
            "SubassuntoOriginal": _clean(source.get("Subassunto - Descrição")),
            "UsuarioAtualNome": internal,
            "SituacaoOriginal": _clean(source.get("Situação")),
            "SetorAtualFonte": _clean(source.get("Centro de Custo Atual - Descrição")),
            "PessoaResponsavelExterna": technical or requester,
            "TipoPessoaResponsavel": "Responsável/RT" if technical else ("Requerente" if requester else ""),
        })
        records[protocol_id] = record
    return records


def _parse_indicator_base(workbook) -> dict[str, dict[str, str]]:
    sheet = workbook["01_RECEBIDOS"]
    records: dict[str, dict[str, str]] = {}
    required = {"ProtocoloID", "NomeRequerente", "ResponsavelTecnico", "ObservacaoUltimoTramite"}
    for source in _sheet_records(sheet, required):
        protocol_id = _protocol_id(source.get("ProtocoloID"))
        if not protocol_id:
            continue
        if protocol_id in records:
            raise ValueError("A planilha possui protocolos duplicados.")
        record = _blank_record(protocol_id)
        record.update({
            "NomeRequerente": _clean(source.get("NomeRequerente")),
            "ResponsavelTecnico": _clean(source.get("ResponsavelTecnico")),
            "ObservacaoUltimoTramite": _clean(source.get("ObservacaoUltimoTramite")),
            "PessoaResponsavelExterna": _clean(source.get("PessoaResponsavelExterna")),
            "TipoPessoaResponsavel": _clean(source.get("TipoPessoaResponsavel")),
        })
        records[protocol_id] = record

    if "03_ESTOQUE" in workbook.sheetnames:
        stock = workbook["03_ESTOQUE"]
        try:
            stock_rows = _sheet_records(stock, {"ProtocoloID", "ResponsavelInterno"})
            for source in stock_rows:
                protocol_id = _protocol_id(source.get("ProtocoloID"))
                if protocol_id in records:
                    records[protocol_id]["ResponsavelInterno"] = _clean(source.get("ResponsavelInterno"))
        except ValueError:
            pass
    return records


def parse_private_xlsx(body: bytes) -> tuple[str, dict[str, dict[str, str]]]:
    if not body:
        raise ValueError("Arquivo vazio.")
    try:
        workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("O arquivo não é uma planilha XLSX/XLSM válida.") from exc

    try:
        if "BASE_COMPLETA" in workbook.sheetnames:
            source_kind = "BASE_COMPLETA"
            records = _parse_complete_base(workbook["BASE_COMPLETA"])
        elif "Report" in workbook.sheetnames:
            source_kind = "IPM_REPORT"
            records = _parse_ipm_report(workbook["Report"])
        elif "BASE23-26" in workbook.sheetnames:
            source_kind = "BASE23-26"
            records = _parse_raw_base(workbook["BASE23-26"])
        elif "01_RECEBIDOS" in workbook.sheetnames:
            source_kind = "BASE_INDICADORES"
            records = _parse_indicator_base(workbook)
        else:
            raise ValueError(
                "Planilha incompatível. Use o relatório IPM com aba Report, a BASE_COMPLETA, a base com aba BASE23-26 ou a BASE_INDICADORES com aba 01_RECEBIDOS."
            )
    finally:
        workbook.close()

    if not records:
        raise ValueError("Nenhum protocolo privado válido foi encontrado na planilha.")
    return source_kind, records


def _persist_private_payload(payload: dict[str, Any]) -> None:
    if not os.getenv("BLOB_READ_WRITE_TOKEN"):
        raise RuntimeError("Armazenamento privado não configurado no servidor.")
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    compressed = lzma.compress(encoded, preset=6)
    try:
        from vercel.blob import BlobClient
        with BlobClient() as client:
            client.put(
                BLOB_PATH,
                compressed,
                access="private",
                content_type="application/x-xz",
                overwrite=True,
            )
    except Exception as exc:
        raise RuntimeError("Não foi possível gravar a base no armazenamento privado.") from exc


def install_private_xlsx(body: bytes, source_name: str = "") -> dict[str, Any]:
    source_kind, parsed = parse_private_xlsx(body)
    public_rows = [dict(row) for row in load_rows()]
    public_ids = {_clean(row.get("ProtocoloID")) for row in public_rows if _clean(row.get("ProtocoloID"))}
    matched_ids = public_ids.intersection(parsed)
    coverage = len(matched_ids) / len(public_ids) if public_ids else 0.0
    if coverage < MIN_PUBLIC_COVERAGE:
        raise ValueError(
            f"A planilha não corresponde à base atual: cobertura de {coverage * 100:.1f}% (mínimo {MIN_PUBLIC_COVERAGE * 100:.0f}%)."
        )

    # Guarda todos os protocolos 2025+ presentes na fonte, inclusive os que ainda
    # não chegaram ao payload público. Assim a mesma fonte pode ser sincronizada
    # antes do deploy e continuará válida depois que os novos protocolos entrarem.
    records = [parsed[protocol_id] for protocol_id in sorted(parsed)]
    installed_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    payload = {
        "v": 3,
        "fields": sorted(PRIVATE_FIELDS_V3),
        "records": records,
        "meta": {
            "installed_at": installed_at,
            "source_kind": source_kind,
            "source_name": _clean(source_name)[:180],
            "public_rows_at_sync": len(public_ids),
            "source_scope_rows": len(parsed),
            "matched_rows_at_sync": len(matched_ids),
            "missing_public_rows_at_sync": len(public_ids) - len(matched_ids),
        },
    }
    _persist_private_payload(payload)
    load_private_rows.cache_clear()

    return {
        "ok": True,
        "configured": True,
        "source_kind": source_kind,
        "public_rows": len(public_ids),
        "private_rows": len(records),
        "matched_rows": len(matched_ids),
        "missing_private_rows": len(public_ids) - len(matched_ids),
        "coverage_percent": round(coverage * 100, 2),
        "installed_at": installed_at,
    }


__all__ = ["install_private_xlsx", "parse_private_xlsx"]