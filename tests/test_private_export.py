import gzip
import json
import os
import tempfile
import unittest
from unittest.mock import patch
from io import BytesIO

from openpyxl import load_workbook

from backend.private_data import PRIVATE_FIELDS_V3, load_private_rows
from backend.private_export import build_private_xlsx


class PrivateExportTests(unittest.TestCase):
    def tearDown(self):
        load_private_rows.cache_clear()

    def test_private_v3_contract_loads_without_publishing_file(self):
        record = {field: "" for field in PRIVATE_FIELDS_V3}
        record.update({
            "ProtocoloID": "2026-1",
            "NomeRequerente": "Pessoa Teste",
            "ObservacaoAbertura": "Solicitação de teste",
            "ObservacaoUltimoTramite": "Trâmite de teste",
        })
        payload = {"v": 3, "fields": sorted(PRIVATE_FIELDS_V3), "records": [record]}
        # No Windows, NamedTemporaryFile mantém um bloqueio exclusivo enquanto o
        # contexto está aberto. O carregador precisa reabrir o artefato pelo
        # caminho, portanto fechamos o descritor antes da leitura.
        with tempfile.NamedTemporaryFile(suffix=".json.gz", delete=False) as tmp:
            tmp.write(gzip.compress(json.dumps(payload, ensure_ascii=False).encode("utf-8")))
            temporary_path = tmp.name
        try:
            load_private_rows.cache_clear()
            with patch.dict(os.environ, {"SEPLANBI_PRIVATE_DATA_PATH": temporary_path}, clear=False):
                rows = load_private_rows()
        finally:
            os.unlink(temporary_path)
        self.assertEqual(rows["2026-1"]["NomeRequerente"], "Pessoa Teste")
        self.assertEqual(rows["2026-1"]["ObservacaoAbertura"], "Solicitação de teste")

    @patch("backend.private_export.core.metadata")
    @patch("backend.private_export.load_private_rows")
    @patch("backend.private_export.load_rows")
    def test_xlsx_joins_public_and_private_fields(self, public_rows, private_rows, metadata):
        public_rows.return_value = [{
            "ProtocoloID": "2026-1",
            "NumeroAnoOriginal": "1/2026",
            "ProtocoloAno": 2026,
            "DataAbertura": "2026-01-02",
            "Macroprocesso": "Licenciamento de Obras",
            "Categoria": "Alvará de Construção",
            "StatusOperacional": "Em Análise",
        }]
        private_rows.return_value = {"2026-1": {
            "ProtocoloID": "2026-1",
            "NomeRequerente": "Pessoa Teste",
            "ResponsavelTecnico": "RT Teste",
            "ResponsavelInterno": "Servidor Teste",
            "PessoaResponsavelExterna": "Pessoa Teste",
            "TipoPessoaResponsavel": "Requerente",
            "ObservacaoUltimoTramite": "Último trâmite",
            "ObservacaoAbertura": "Observação de abertura",
            "SubassuntoOriginal": "ALVARÁ",
            "UsuarioAtualNome": "Usuário Teste",
            "SituacaoOriginal": "Aberto",
            "SetorAtualFonte": "Engenharia",
        }}
        metadata.return_value = {"source_updated_at": "2026-08-29", "taxonomy_version": "V07"}

        body, filename = build_private_xlsx()
        self.assertTrue(filename.endswith("20260829.xlsx"))
        workbook = load_workbook(BytesIO(body), read_only=True)
        sheet = workbook["BASE_COMPLETA"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        values = dict(zip(headers, row))
        self.assertEqual(values["NomeRequerente"], "Pessoa Teste")
        self.assertEqual(values["ObservacaoAbertura"], "Observação de abertura")
        self.assertEqual(values["ObservacaoUltimoTramite"], "Último trâmite")
        self.assertEqual(values["StatusOperacional"], "Em Análise")


if __name__ == "__main__":
    unittest.main()
